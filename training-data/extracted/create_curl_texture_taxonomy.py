"""
Comprehensive Curl Pattern, Texture, and Hair Properties Taxonomy Generator
Creates detailed Excel files with full spectrum classifications
"""

import pandas as pd
from openpyxl.utils import get_column_letter

# =============================================================================
# CURL PATTERN TAXONOMY (Andre Walker System - Detailed)
# =============================================================================

curl_patterns = [
    # TYPE 1 - STRAIGHT
    {
        "Type": "1A",
        "Category": "Straight",
        "Sub_Category": "Fine Straight",
        "Description": "Completely straight, very fine, soft, shiny, difficult to hold curl",
        "Strand_Diameter": "Fine (< 60 micrometers)",
        "Visual_Characteristics": "Lies flat against scalp, reflects maximum light, no wave pattern",
        "Common_Ethnicities": "Asian, Northern European",
        "Porosity_Tendency": "Low",
        "Volume_Tendency": "Low - tends to be flat",
        "Frizz_Tendency": "Very Low",
        "Styling_Challenges": "Difficult to add volume, curls fall out quickly",
        "Recommended_Products": "Volumizing mousse, light hold sprays",
        "Curl_Circumference": "N/A - No curl",
        "S_Pattern": "None",
        "Shrinkage": "0%"
    },
    {
        "Type": "1B",
        "Category": "Straight",
        "Sub_Category": "Medium Straight",
        "Description": "Straight with medium texture, more body than 1A, slight bend possible",
        "Strand_Diameter": "Medium (60-80 micrometers)",
        "Visual_Characteristics": "Straight with slight body, some natural movement",
        "Common_Ethnicities": "European, Asian, Mixed",
        "Porosity_Tendency": "Low-Medium",
        "Volume_Tendency": "Medium",
        "Frizz_Tendency": "Low",
        "Styling_Challenges": "Can become oily at roots quickly",
        "Recommended_Products": "Lightweight serums, dry shampoo",
        "Curl_Circumference": "N/A - No curl",
        "S_Pattern": "None to very slight",
        "Shrinkage": "0%"
    },
    {
        "Type": "1C",
        "Category": "Straight",
        "Sub_Category": "Coarse Straight",
        "Description": "Straight and coarse, resistant to styling, can have slight wave",
        "Strand_Diameter": "Coarse (> 80 micrometers)",
        "Visual_Characteristics": "Thick individual strands, may have subtle wave, less shine",
        "Common_Ethnicities": "Asian, Mediterranean, Middle Eastern",
        "Porosity_Tendency": "Low",
        "Volume_Tendency": "High",
        "Frizz_Tendency": "Low-Medium",
        "Styling_Challenges": "Resistant to curling, can be wiry",
        "Recommended_Products": "Smoothing serums, anti-humidity products",
        "Curl_Circumference": "N/A - Minimal to no curl",
        "S_Pattern": "None to very slight",
        "Shrinkage": "0-5%"
    },
    
    # TYPE 2 - WAVY
    {
        "Type": "2A",
        "Category": "Wavy",
        "Sub_Category": "Loose Waves",
        "Description": "Loose, stretched S-pattern waves, fine texture, easily straightened",
        "Strand_Diameter": "Fine to Medium",
        "Visual_Characteristics": "Gentle S-bend, waves start mid-length, beachy texture",
        "Common_Ethnicities": "European, Latin American, Mixed",
        "Porosity_Tendency": "Low-Medium",
        "Volume_Tendency": "Medium",
        "Frizz_Tendency": "Low-Medium",
        "Styling_Challenges": "Waves can fall flat, inconsistent pattern",
        "Recommended_Products": "Sea salt spray, light mousse, wave enhancers",
        "Curl_Circumference": "Large (> 2 inches)",
        "S_Pattern": "Loose S-shape",
        "Shrinkage": "5-10%"
    },
    {
        "Type": "2B",
        "Category": "Wavy",
        "Sub_Category": "Defined Waves",
        "Description": "More defined S-pattern, waves from root, medium texture",
        "Strand_Diameter": "Medium",
        "Visual_Characteristics": "Defined S-waves throughout, more consistent pattern",
        "Common_Ethnicities": "European, Mediterranean, Latin American",
        "Porosity_Tendency": "Medium",
        "Volume_Tendency": "Medium-High",
        "Frizz_Tendency": "Medium",
        "Styling_Challenges": "Frizz at crown, can be unpredictable",
        "Recommended_Products": "Curl cream, anti-frizz serum, diffuser",
        "Curl_Circumference": "Medium-Large (1.5-2 inches)",
        "S_Pattern": "Defined S-shape",
        "Shrinkage": "10-15%"
    },
    {
        "Type": "2C",
        "Category": "Wavy",
        "Sub_Category": "Strong Waves",
        "Description": "Strong S-waves verging on loose curls, coarser texture, prone to frizz",
        "Strand_Diameter": "Medium to Coarse",
        "Visual_Characteristics": "Thick waves, some spiral formation, high volume",
        "Common_Ethnicities": "Mediterranean, Middle Eastern, Mixed",
        "Porosity_Tendency": "Medium-High",
        "Volume_Tendency": "High",
        "Frizz_Tendency": "High",
        "Styling_Challenges": "Frizz control, definition maintenance",
        "Recommended_Products": "Leave-in conditioner, curl defining cream, gel",
        "Curl_Circumference": "Medium (1-1.5 inches)",
        "S_Pattern": "Strong S to loose spiral",
        "Shrinkage": "15-25%"
    },
    
    # TYPE 3 - CURLY
    {
        "Type": "3A",
        "Category": "Curly",
        "Sub_Category": "Loose Curls",
        "Description": "Large, loose spiral curls, S-shaped, well-defined, shiny",
        "Strand_Diameter": "Fine to Medium",
        "Visual_Characteristics": "Big loopy curls, sidewalk chalk circumference, bouncy",
        "Common_Ethnicities": "Mixed, Latin American, Mediterranean, African descent",
        "Porosity_Tendency": "Medium",
        "Volume_Tendency": "High",
        "Frizz_Tendency": "Medium",
        "Styling_Challenges": "Weight can pull curls flat, needs moisture",
        "Recommended_Products": "Curl cream, lightweight gel, leave-in conditioner",
        "Curl_Circumference": "Large spiral (sidewalk chalk size)",
        "S_Pattern": "Defined spiral/ringlet",
        "Shrinkage": "25-35%"
    },
    {
        "Type": "3B",
        "Category": "Curly",
        "Sub_Category": "Tight Curls",
        "Description": "Springy, tight curls, marker-sized circumference, voluminous",
        "Strand_Diameter": "Medium",
        "Visual_Characteristics": "Bouncy ringlets, well-defined spirals, high volume",
        "Common_Ethnicities": "African descent, Mixed, Latin American",
        "Porosity_Tendency": "Medium-High",
        "Volume_Tendency": "Very High",
        "Frizz_Tendency": "Medium-High",
        "Styling_Challenges": "Dryness, shrinkage, frizz control",
        "Recommended_Products": "Rich curl cream, strong hold gel, deep conditioner",
        "Curl_Circumference": "Medium spiral (Sharpie marker size)",
        "S_Pattern": "Tight spiral/ringlet",
        "Shrinkage": "35-50%"
    },
    {
        "Type": "3C",
        "Category": "Curly",
        "Sub_Category": "Corkscrew Curls",
        "Description": "Tight corkscrew curls, pencil-sized, densely packed, high shrinkage",
        "Strand_Diameter": "Fine to Medium",
        "Visual_Characteristics": "Tight corkscrews, very dense, straw/pencil circumference",
        "Common_Ethnicities": "African descent, Mixed",
        "Porosity_Tendency": "High",
        "Volume_Tendency": "Very High",
        "Frizz_Tendency": "High",
        "Styling_Challenges": "Extreme shrinkage, dryness, tangles",
        "Recommended_Products": "Heavy cream, butter-based products, LOC method",
        "Curl_Circumference": "Small spiral (pencil/straw size)",
        "S_Pattern": "Tight corkscrew",
        "Shrinkage": "50-65%"
    },
    
    # TYPE 4 - COILY/KINKY
    {
        "Type": "4A",
        "Category": "Coily",
        "Sub_Category": "Soft Coils",
        "Description": "Tightly coiled S-pattern, defined coils, crochet needle circumference",
        "Strand_Diameter": "Fine to Medium",
        "Visual_Characteristics": "Defined S-coils, springy, visible curl pattern when wet",
        "Common_Ethnicities": "African descent",
        "Porosity_Tendency": "High",
        "Volume_Tendency": "Very High",
        "Frizz_Tendency": "High",
        "Styling_Challenges": "Dryness, shrinkage, single strand knots",
        "Recommended_Products": "Heavy butters, oils, LOC/LCO method, twist-outs",
        "Curl_Circumference": "Tight coil (crochet needle size)",
        "S_Pattern": "Tight S-coil",
        "Shrinkage": "65-75%"
    },
    {
        "Type": "4B",
        "Category": "Coily",
        "Sub_Category": "Z-Pattern Coils",
        "Description": "Z-shaped pattern rather than S, sharp angles, less defined coils",
        "Strand_Diameter": "Fine to Medium",
        "Visual_Characteristics": "Zigzag pattern, bends at sharp angles, cotton-like texture",
        "Common_Ethnicities": "African descent",
        "Porosity_Tendency": "High",
        "Volume_Tendency": "Very High",
        "Frizz_Tendency": "Very High",
        "Styling_Challenges": "Extreme dryness, breakage, shrinkage",
        "Recommended_Products": "Heavy creams, butters, protective styles, deep conditioning",
        "Curl_Circumference": "Z-bend (no circular coil)",
        "S_Pattern": "Z-pattern/zigzag",
        "Shrinkage": "70-80%"
    },
    {
        "Type": "4C",
        "Category": "Coily",
        "Sub_Category": "Tight Coils",
        "Description": "Tightest coil pattern, minimal definition, Z-pattern, very fragile",
        "Strand_Diameter": "Fine",
        "Visual_Characteristics": "Very tight zigzag, appears patternless when dry, cotton-like",
        "Common_Ethnicities": "African descent",
        "Porosity_Tendency": "High",
        "Volume_Tendency": "Very High (when stretched)",
        "Frizz_Tendency": "Very High",
        "Styling_Challenges": "Extreme shrinkage (up to 75%), dryness, breakage, tangles",
        "Recommended_Products": "Heavy butters, oils, protective styles, LOC method, satin bonnet",
        "Curl_Circumference": "Tightest Z-bend",
        "S_Pattern": "Tight Z-pattern",
        "Shrinkage": "75-85%"
    }
]

# =============================================================================
# LOIS TEXTURE SYSTEM
# =============================================================================

lois_textures = [
    {
        "Letter": "L",
        "Texture_Name": "L-Pattern (Bent)",
        "Description": "Hair strands have bends with no curves, like the letter L",
        "Visual_Characteristics": "Angular bends, minimal curl definition",
        "Strand_Shape": "Angular/bent",
        "Shine_Level": "Low to Medium",
        "Common_In_Types": "4A, 4B, 4C",
        "Characteristics": "Bends at sharp angles, can appear zigzag"
    },
    {
        "Letter": "O",
        "Texture_Name": "O-Pattern (Coiled)",
        "Description": "Hair strands curl into tight O-shapes or coils",
        "Visual_Characteristics": "Circular coils, spring-like",
        "Strand_Shape": "Circular/coiled",
        "Shine_Level": "Low",
        "Common_In_Types": "3C, 4A, 4B",
        "Characteristics": "Tight coils, high shrinkage"
    },
    {
        "Letter": "I",
        "Texture_Name": "I-Pattern (Straight)",
        "Description": "Hair strands are straight with no bend or curl",
        "Visual_Characteristics": "Completely straight, no wave",
        "Strand_Shape": "Straight/linear",
        "Shine_Level": "High",
        "Common_In_Types": "1A, 1B, 1C",
        "Characteristics": "Lies flat, maximum shine"
    },
    {
        "Letter": "S",
        "Texture_Name": "S-Pattern (Wavy/Curly)",
        "Description": "Hair strands form S-shaped waves or curls",
        "Visual_Characteristics": "Flowing S-curves, waves or spirals",
        "Strand_Shape": "S-curved",
        "Shine_Level": "Medium to High",
        "Common_In_Types": "2A-3B",
        "Characteristics": "Defined waves or curls"
    }
]

# =============================================================================
# STRAND THICKNESS/DIAMETER
# =============================================================================

strand_thickness = [
    {
        "Category": "Fine",
        "Diameter_Range": "< 60 micrometers",
        "Description": "Thin individual strands, often appears less dense",
        "Visual_Test": "Barely visible, feels like silk thread",
        "Touch_Test": "Difficult to feel between fingers",
        "Float_Test": "Floats on water surface",
        "Characteristics": "Easily damaged, prone to breakage, gets oily quickly",
        "Volume_Tendency": "Low natural volume",
        "Product_Needs": "Lightweight products, volumizing formulas",
        "Heat_Sensitivity": "High - use lower temperatures",
        "Common_With_Types": "1A, 3A, 4C"
    },
    {
        "Category": "Medium",
        "Diameter_Range": "60-80 micrometers",
        "Description": "Average strand thickness, most common",
        "Visual_Test": "Visible, similar to sewing thread",
        "Touch_Test": "Can feel between fingers",
        "Float_Test": "Slowly sinks in water",
        "Characteristics": "Versatile, holds styles well, moderate strength",
        "Volume_Tendency": "Medium natural volume",
        "Product_Needs": "Most products work well",
        "Heat_Sensitivity": "Medium - standard temperatures",
        "Common_With_Types": "1B, 2B, 3B"
    },
    {
        "Category": "Coarse",
        "Diameter_Range": "> 80 micrometers",
        "Description": "Thick individual strands, often appears very full",
        "Visual_Test": "Clearly visible, thicker than thread",
        "Touch_Test": "Easily felt, may feel wiry",
        "Float_Test": "Sinks quickly in water",
        "Characteristics": "Strong, resistant to damage, can be dry, resistant to styling",
        "Volume_Tendency": "High natural volume",
        "Product_Needs": "Rich, heavy products, oils and butters",
        "Heat_Sensitivity": "Low - can handle higher temperatures",
        "Common_With_Types": "1C, 2C, 4B"
    }
]

# =============================================================================
# HAIR DENSITY
# =============================================================================

hair_density = [
    {
        "Category": "Low Density",
        "Strands_Per_Sq_Inch": "< 100,000 total",
        "Description": "Fewer hair strands on the scalp",
        "Visual_Test": "Scalp easily visible through hair",
        "Ponytail_Test": "Ponytail circumference < 2 inches",
        "Part_Test": "Wide part, scalp very visible",
        "Characteristics": "Hair may appear thin, scalp shows through",
        "Styling_Considerations": "Needs volumizing techniques, avoid heavy products",
        "Product_Recommendations": "Volumizing mousse, root lifters, dry shampoo"
    },
    {
        "Category": "Medium Density",
        "Strands_Per_Sq_Inch": "100,000 - 150,000 total",
        "Description": "Average number of hair strands",
        "Visual_Test": "Some scalp visible when hair is still",
        "Ponytail_Test": "Ponytail circumference 2-3 inches",
        "Part_Test": "Moderate part width",
        "Characteristics": "Balanced fullness, versatile styling",
        "Styling_Considerations": "Most styles work well",
        "Product_Recommendations": "Standard products work well"
    },
    {
        "Category": "High Density",
        "Strands_Per_Sq_Inch": "> 150,000 total",
        "Description": "Many hair strands packed closely",
        "Visual_Test": "Scalp barely visible or not visible",
        "Ponytail_Test": "Ponytail circumference > 3 inches",
        "Part_Test": "Narrow part, scalp hidden",
        "Characteristics": "Very full appearance, can be heavy, takes longer to dry",
        "Styling_Considerations": "May need thinning, layers help with weight",
        "Product_Recommendations": "Smoothing products, may need more product quantity"
    }
]

# =============================================================================
# POROSITY
# =============================================================================

porosity = [
    {
        "Category": "Low Porosity",
        "Description": "Cuticles are tightly closed, resistant to moisture",
        "Float_Test": "Hair floats on water for extended time",
        "Spray_Test": "Water beads up on hair surface",
        "Characteristics": "Product buildup prone, takes long to dry, resistant to color",
        "Moisture_Behavior": "Difficult to absorb moisture, but retains it once absorbed",
        "Heat_Response": "Heat helps open cuticles for product absorption",
        "Product_Needs": "Lightweight products, heat for deep conditioning, clarifying shampoo",
        "Best_Ingredients": "Humectants (glycerin, honey), lightweight oils",
        "Avoid": "Heavy butters, protein overload"
    },
    {
        "Category": "Medium/Normal Porosity",
        "Description": "Cuticles are slightly raised, balanced moisture absorption",
        "Float_Test": "Hair floats then slowly sinks",
        "Spray_Test": "Water absorbs at moderate rate",
        "Characteristics": "Holds styles well, accepts color easily, balanced moisture",
        "Moisture_Behavior": "Absorbs and retains moisture well",
        "Heat_Response": "Responds well to heat styling",
        "Product_Needs": "Balanced products, occasional deep conditioning",
        "Best_Ingredients": "Most ingredients work well",
        "Avoid": "Over-processing"
    },
    {
        "Category": "High Porosity",
        "Description": "Cuticles are raised or damaged, absorbs moisture quickly but loses it fast",
        "Float_Test": "Hair sinks immediately",
        "Spray_Test": "Water absorbs instantly",
        "Characteristics": "Frizz prone, dries quickly, tangles easily, color fades fast",
        "Moisture_Behavior": "Absorbs moisture quickly but cannot retain it",
        "Heat_Response": "Very susceptible to heat damage",
        "Product_Needs": "Heavy products, protein treatments, sealants",
        "Best_Ingredients": "Proteins, heavy butters, oils to seal",
        "Avoid": "Excessive heat, harsh chemicals"
    }
]

# =============================================================================
# HAIR LENGTH CLASSIFICATIONS
# =============================================================================

hair_length = [
    {
        "Category": "Buzzed/Shaved",
        "Length_Range": "< 0.5 inches",
        "Description": "Very short, close to scalp",
        "Reference_Point": "Does not extend past scalp",
        "Styling_Options": "Limited, focus on scalp care",
        "Growth_Time_From_Bald": "0-2 months"
    },
    {
        "Category": "TWA (Teeny Weeny Afro)",
        "Length_Range": "0.5 - 2 inches",
        "Description": "Short natural hair, common after big chop",
        "Reference_Point": "Above ears, does not cover forehead",
        "Styling_Options": "Finger coils, wash and go, TWA styles",
        "Growth_Time_From_Bald": "2-6 months"
    },
    {
        "Category": "Short",
        "Length_Range": "2 - 4 inches",
        "Description": "Above chin length",
        "Reference_Point": "Ear length to above chin",
        "Styling_Options": "Pixie cuts, short bobs, defined curls",
        "Growth_Time_From_Bald": "6-12 months"
    },
    {
        "Category": "Chin Length",
        "Length_Range": "4 - 6 inches",
        "Description": "Hair reaches chin",
        "Reference_Point": "At chin level",
        "Styling_Options": "Bobs, lobs, half-up styles",
        "Growth_Time_From_Bald": "12-18 months"
    },
    {
        "Category": "Shoulder Length",
        "Length_Range": "6 - 10 inches",
        "Description": "Hair reaches shoulders",
        "Reference_Point": "Touches shoulders when straight",
        "Styling_Options": "Most styles possible, ponytails, braids",
        "Growth_Time_From_Bald": "18-30 months"
    },
    {
        "Category": "Armpit Length (APL)",
        "Length_Range": "10 - 14 inches",
        "Description": "Hair reaches armpit level",
        "Reference_Point": "Armpit when arm is down",
        "Styling_Options": "Full range of styles, updos, long braids",
        "Growth_Time_From_Bald": "30-42 months"
    },
    {
        "Category": "Bra Strap Length (BSL)",
        "Length_Range": "14 - 18 inches",
        "Description": "Hair reaches bra strap",
        "Reference_Point": "Middle of back at bra strap",
        "Styling_Options": "All styles, dramatic updos",
        "Growth_Time_From_Bald": "42-54 months"
    },
    {
        "Category": "Mid-Back Length (MBL)",
        "Length_Range": "18 - 22 inches",
        "Description": "Hair reaches middle of back",
        "Reference_Point": "Between bra strap and waist",
        "Styling_Options": "All styles, weight may affect curls",
        "Growth_Time_From_Bald": "54-66 months"
    },
    {
        "Category": "Waist Length",
        "Length_Range": "22 - 26 inches",
        "Description": "Hair reaches waist",
        "Reference_Point": "Natural waist level",
        "Styling_Options": "All styles, may need layers for movement",
        "Growth_Time_From_Bald": "66-78 months"
    },
    {
        "Category": "Hip Length",
        "Length_Range": "26 - 32 inches",
        "Description": "Hair reaches hips",
        "Reference_Point": "Hip bone level",
        "Styling_Options": "Dramatic length, special care needed",
        "Growth_Time_From_Bald": "78-96 months"
    },
    {
        "Category": "Tailbone Length",
        "Length_Range": "32 - 38 inches",
        "Description": "Hair reaches tailbone",
        "Reference_Point": "Tailbone/coccyx level",
        "Styling_Options": "Very long, high maintenance",
        "Growth_Time_From_Bald": "96-114 months"
    },
    {
        "Category": "Classic Length",
        "Length_Range": "38+ inches",
        "Description": "Hair reaches below tailbone to thighs",
        "Reference_Point": "Below tailbone",
        "Styling_Options": "Extremely long, requires dedicated care",
        "Growth_Time_From_Bald": "114+ months"
    }
]

# =============================================================================
# HAIR HEALTH INDICATORS
# =============================================================================

health_indicators = [
    # CUTICLE CONDITION
    {"Category": "Cuticle Condition", "Level": "Healthy/Smooth", "Description": "Cuticles lie flat, hair is shiny and smooth", "Visual_Signs": "High shine, smooth to touch, no tangles", "Causes": "Proper care, minimal damage", "Treatment": "Maintain current routine"},
    {"Category": "Cuticle Condition", "Level": "Slightly Raised", "Description": "Cuticles slightly lifted, some roughness", "Visual_Signs": "Reduced shine, slight roughness", "Causes": "Minor damage, environmental factors", "Treatment": "Deep conditioning, protein treatments"},
    {"Category": "Cuticle Condition", "Level": "Raised", "Description": "Cuticles significantly lifted, porous", "Visual_Signs": "Dull appearance, tangles easily, frizzy", "Causes": "Chemical damage, heat damage", "Treatment": "Protein treatments, bond repair"},
    {"Category": "Cuticle Condition", "Level": "Damaged/Missing", "Description": "Cuticles severely damaged or missing", "Visual_Signs": "Very dull, gummy when wet, breaks easily", "Causes": "Severe chemical/heat damage", "Treatment": "Trim, intensive repair, protective styles"},
    
    # BREAKAGE LEVELS
    {"Category": "Breakage", "Level": "None", "Description": "No visible breakage", "Visual_Signs": "Even length, no short pieces", "Causes": "Good hair care practices", "Treatment": "Maintain routine"},
    {"Category": "Breakage", "Level": "Mild", "Description": "Occasional short pieces", "Visual_Signs": "Few flyaways, minimal short hairs", "Causes": "Minor mechanical damage", "Treatment": "Gentle handling, silk pillowcase"},
    {"Category": "Breakage", "Level": "Moderate", "Description": "Noticeable breakage throughout", "Visual_Signs": "Many short pieces, uneven texture", "Causes": "Damage accumulation", "Treatment": "Protein treatments, reduce manipulation"},
    {"Category": "Breakage", "Level": "Severe", "Description": "Significant breakage, hair not retaining length", "Visual_Signs": "Lots of short pieces, thinning", "Causes": "Severe damage, health issues", "Treatment": "Professional help, possible trim, medical check"},
    
    # SPLIT ENDS
    {"Category": "Split Ends", "Level": "None", "Description": "No split ends visible", "Visual_Signs": "Smooth, tapered ends", "Causes": "Regular trims, good care", "Treatment": "Maintain routine"},
    {"Category": "Split Ends", "Level": "Mild", "Description": "Few split ends", "Visual_Signs": "Occasional Y-shaped splits", "Causes": "Normal wear", "Treatment": "Trim, use serums"},
    {"Category": "Split Ends", "Level": "Moderate", "Description": "Multiple split ends visible", "Visual_Signs": "Many splits, some traveling up shaft", "Causes": "Lack of trims, damage", "Treatment": "Trim 1-2 inches, deep condition"},
    {"Category": "Split Ends", "Level": "Severe", "Description": "Extensive splitting", "Visual_Signs": "Tree-like splits, feathered ends, white dots", "Causes": "Severe neglect or damage", "Treatment": "Significant trim needed"},
    
    # HYDRATION
    {"Category": "Hydration", "Level": "Well-Moisturized", "Description": "Optimal moisture balance", "Visual_Signs": "Soft, supple, defined curls, elastic", "Causes": "Good moisture routine", "Treatment": "Maintain routine"},
    {"Category": "Hydration", "Level": "Slightly Dry", "Description": "Needs more moisture", "Visual_Signs": "Slightly rough, less defined", "Causes": "Environmental factors, product buildup", "Treatment": "Deep condition, add leave-in"},
    {"Category": "Hydration", "Level": "Dry", "Description": "Lacking moisture", "Visual_Signs": "Rough, dull, tangles, undefined", "Causes": "Insufficient moisture routine", "Treatment": "LOC/LCO method, deep conditioning"},
    {"Category": "Hydration", "Level": "Very Dry/Brittle", "Description": "Severely dehydrated", "Visual_Signs": "Straw-like, snaps easily, no elasticity", "Causes": "Severe neglect, damage, health issues", "Treatment": "Intensive moisture treatments, check health"},
    
    # ELASTICITY
    {"Category": "Elasticity", "Level": "High/Normal", "Description": "Hair stretches and returns to shape", "Visual_Signs": "Bouncy curls, no breakage when stretched", "Causes": "Balanced protein/moisture", "Treatment": "Maintain balance"},
    {"Category": "Elasticity", "Level": "Low (Protein Deficient)", "Description": "Hair stretches but doesn't return, mushy", "Visual_Signs": "Limp, overly soft, stretches far", "Causes": "Too much moisture, not enough protein", "Treatment": "Protein treatment"},
    {"Category": "Elasticity", "Level": "Low (Moisture Deficient)", "Description": "Hair snaps without stretching", "Visual_Signs": "Brittle, snaps immediately", "Causes": "Too much protein, not enough moisture", "Treatment": "Deep moisture treatment"},
]

# =============================================================================
# SCALP CONDITIONS
# =============================================================================

scalp_conditions = [
    {"Condition": "Normal/Healthy", "Description": "Balanced oil production, no flaking or irritation", "Visual_Signs": "Clear scalp, no visible issues", "Causes": "Good scalp care", "Treatment": "Maintain routine"},
    {"Condition": "Dry", "Description": "Insufficient sebum production", "Visual_Signs": "Tight feeling, small white flakes, itchy", "Causes": "Over-washing, harsh products, climate", "Treatment": "Scalp oils, less frequent washing, gentle products"},
    {"Condition": "Oily", "Description": "Excess sebum production", "Visual_Signs": "Greasy appearance, limp roots", "Causes": "Genetics, hormones, over-washing", "Treatment": "Balanced washing, clarifying shampoo, avoid heavy products at roots"},
    {"Condition": "Combination", "Description": "Oily at roots, dry at ends", "Visual_Signs": "Greasy scalp, dry lengths", "Causes": "Common in longer hair", "Treatment": "Targeted treatment - clarify roots, moisturize ends"},
    {"Condition": "Flaky (Dandruff)", "Description": "Visible flaking, may be itchy", "Visual_Signs": "White or yellow flakes, itching", "Causes": "Malassezia fungus, dry scalp, product buildup", "Treatment": "Anti-dandruff shampoo, scalp treatments"},
    {"Condition": "Sensitive/Irritated", "Description": "Easily irritated, reactive", "Visual_Signs": "Redness, burning, itching", "Causes": "Product sensitivity, conditions like eczema", "Treatment": "Fragrance-free products, gentle formulas, dermatologist"},
    {"Condition": "Product Buildup", "Description": "Accumulation of product residue", "Visual_Signs": "Dull hair, flaky residue, limp", "Causes": "Insufficient cleansing, heavy products", "Treatment": "Clarifying shampoo, ACV rinse"},
]

# =============================================================================
# STYLING STATES
# =============================================================================

styling_states = [
    {"State": "Natural/Unstyled", "Description": "Hair in its natural state without manipulation", "Characteristics": "True texture visible", "Maintenance": "Basic wash and go"},
    {"State": "Wash and Go", "Description": "Styled wet and left to air dry", "Characteristics": "Natural curl pattern enhanced", "Maintenance": "Refresh with water/product"},
    {"State": "Twist Out", "Description": "Hair twisted while wet, unraveled when dry", "Characteristics": "Defined, elongated curls", "Maintenance": "Pineapple at night, refresh as needed"},
    {"State": "Braid Out", "Description": "Hair braided while wet, unraveled when dry", "Characteristics": "Wavy, crimped pattern", "Maintenance": "Similar to twist out"},
    {"State": "Bantu Knot Out", "Description": "Hair sectioned and knotted, unraveled when dry", "Characteristics": "Defined curls with volume", "Maintenance": "Refresh with light product"},
    {"State": "Flexi Rod/Roller Set", "Description": "Hair set on rods or rollers", "Characteristics": "Uniform curls, bouncy", "Maintenance": "Wrap at night"},
    {"State": "Silk Press/Flat Iron", "Description": "Hair straightened with heat", "Characteristics": "Straight, sleek, shiny", "Maintenance": "Wrap at night, avoid moisture"},
    {"State": "Blowout", "Description": "Hair dried with tension using blow dryer", "Characteristics": "Stretched, voluminous, not fully straight", "Maintenance": "Can last several days"},
    {"State": "Protective Style - Braids", "Description": "Hair braided close to scalp or in extensions", "Characteristics": "Low manipulation, hair protected", "Maintenance": "Scalp care, moisturize, 4-8 weeks max"},
    {"State": "Protective Style - Twists", "Description": "Hair twisted, may include extensions", "Characteristics": "Low manipulation", "Maintenance": "Similar to braids"},
    {"State": "Protective Style - Locs", "Description": "Hair permanently locked/matted", "Characteristics": "Permanent style, various stages", "Maintenance": "Regular retwisting, washing"},
    {"State": "Wig/Weave", "Description": "Additional hair added for coverage or length", "Characteristics": "Versatile styling, protects natural hair", "Maintenance": "Care for both wig and natural hair underneath"},
    {"State": "Updo", "Description": "Hair gathered and pinned up", "Characteristics": "Hair off neck and shoulders", "Maintenance": "Secure pins, refresh as needed"},
    {"State": "Ponytail/Bun", "Description": "Hair gathered and secured", "Characteristics": "Quick, versatile", "Maintenance": "Avoid tight styles, use satin scrunchies"},
]

# =============================================================================
# CREATE EXCEL FILES
# =============================================================================

def create_curl_texture_excel():
    # Create DataFrames
    df_curl = pd.DataFrame(curl_patterns)
    df_lois = pd.DataFrame(lois_textures)
    df_thickness = pd.DataFrame(strand_thickness)
    df_density = pd.DataFrame(hair_density)
    df_porosity = pd.DataFrame(porosity)
    df_length = pd.DataFrame(hair_length)
    df_health = pd.DataFrame(health_indicators)
    df_scalp = pd.DataFrame(scalp_conditions)
    df_styling = pd.DataFrame(styling_states)
    
    # Write Curl Patterns Excel
    with pd.ExcelWriter('/home/ubuntu/hair_taxonomy/excel/curl_pattern_taxonomy.xlsx', engine='openpyxl') as writer:
        df_curl.to_excel(writer, sheet_name='Curl Patterns (1A-4C)', index=False)
        df_lois.to_excel(writer, sheet_name='LOIS Texture System', index=False)
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Write Hair Properties Excel
    with pd.ExcelWriter('/home/ubuntu/hair_taxonomy/excel/hair_properties_taxonomy.xlsx', engine='openpyxl') as writer:
        df_thickness.to_excel(writer, sheet_name='Strand Thickness', index=False)
        df_density.to_excel(writer, sheet_name='Hair Density', index=False)
        df_porosity.to_excel(writer, sheet_name='Porosity', index=False)
        df_length.to_excel(writer, sheet_name='Hair Length', index=False)
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Write Hair Health & Condition Excel
    with pd.ExcelWriter('/home/ubuntu/hair_taxonomy/excel/hair_health_taxonomy.xlsx', engine='openpyxl') as writer:
        df_health.to_excel(writer, sheet_name='Health Indicators', index=False)
        df_scalp.to_excel(writer, sheet_name='Scalp Conditions', index=False)
        df_styling.to_excel(writer, sheet_name='Styling States', index=False)
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print("Curl Pattern Taxonomy Excel created successfully!")
    print("Hair Properties Taxonomy Excel created successfully!")
    print("Hair Health Taxonomy Excel created successfully!")

if __name__ == "__main__":
    create_curl_texture_excel()
