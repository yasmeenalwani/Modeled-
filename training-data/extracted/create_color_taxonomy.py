"""
Comprehensive Hair Color Taxonomy Generator
Creates detailed Excel file with full spectrum color classifications
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# =============================================================================
# HAIR COLOR LEVEL SYSTEM (International Color Chart - Level 1-10)
# =============================================================================

color_levels = [
    {
        "Level": 1,
        "Level_Name": "Black",
        "Description": "Darkest natural hair color, appears jet black",
        "Underlying_Pigment": "Red",
        "Melanin_Type": "Eumelanin (high)",
        "RGB_Approximate": "(10, 10, 10)",
        "Hex_Code": "#0A0A0A",
        "Common_Names": "Jet Black, Blue-Black, Raven",
        "Percentage_Population": "~75% globally"
    },
    {
        "Level": 2,
        "Level_Name": "Darkest Brown",
        "Description": "Very dark brown, almost black but with brown undertones",
        "Underlying_Pigment": "Red",
        "Melanin_Type": "Eumelanin (high)",
        "RGB_Approximate": "(35, 25, 20)",
        "Hex_Code": "#231914",
        "Common_Names": "Off-Black, Soft Black, Espresso",
        "Percentage_Population": "Common in Asia, Africa, Latin America"
    },
    {
        "Level": 3,
        "Level_Name": "Dark Brown",
        "Description": "Rich dark brown with visible brown tones in light",
        "Underlying_Pigment": "Red-Orange",
        "Melanin_Type": "Eumelanin (medium-high)",
        "RGB_Approximate": "(60, 40, 30)",
        "Hex_Code": "#3C281E",
        "Common_Names": "Dark Chocolate, Mocha, Coffee",
        "Percentage_Population": "Very common globally"
    },
    {
        "Level": 4,
        "Level_Name": "Medium Brown",
        "Description": "True brown, balanced depth",
        "Underlying_Pigment": "Orange",
        "Melanin_Type": "Eumelanin (medium)",
        "RGB_Approximate": "(90, 60, 40)",
        "Hex_Code": "#5A3C28",
        "Common_Names": "Chestnut, Milk Chocolate, Brunette",
        "Percentage_Population": "Common in Europe, Americas"
    },
    {
        "Level": 5,
        "Level_Name": "Light Brown",
        "Description": "Lighter brown with warm undertones",
        "Underlying_Pigment": "Orange-Gold",
        "Melanin_Type": "Mixed eumelanin/pheomelanin",
        "RGB_Approximate": "(120, 80, 55)",
        "Hex_Code": "#785037",
        "Common_Names": "Caramel, Toffee, Light Chestnut",
        "Percentage_Population": "Common in Europe, Americas"
    },
    {
        "Level": 6,
        "Level_Name": "Dark Blonde",
        "Description": "Transition between brown and blonde",
        "Underlying_Pigment": "Gold-Orange",
        "Melanin_Type": "Mixed (more pheomelanin)",
        "RGB_Approximate": "(150, 110, 70)",
        "Hex_Code": "#966E46",
        "Common_Names": "Dirty Blonde, Dark Honey, Bronde",
        "Percentage_Population": "~2% globally"
    },
    {
        "Level": 7,
        "Level_Name": "Medium Blonde",
        "Description": "True blonde with golden undertones",
        "Underlying_Pigment": "Yellow-Gold",
        "Melanin_Type": "Pheomelanin dominant",
        "RGB_Approximate": "(180, 140, 90)",
        "Hex_Code": "#B48C5A",
        "Common_Names": "Golden Blonde, Honey Blonde, Wheat",
        "Percentage_Population": "~2% globally"
    },
    {
        "Level": 8,
        "Level_Name": "Light Blonde",
        "Description": "Light blonde with visible yellow tones",
        "Underlying_Pigment": "Yellow",
        "Melanin_Type": "Pheomelanin (low)",
        "RGB_Approximate": "(210, 175, 120)",
        "Hex_Code": "#D2AF78",
        "Common_Names": "Butter Blonde, Champagne, Sandy Blonde",
        "Percentage_Population": "~1.5% globally"
    },
    {
        "Level": 9,
        "Level_Name": "Very Light Blonde",
        "Description": "Very light with pale yellow undertones",
        "Underlying_Pigment": "Pale Yellow",
        "Melanin_Type": "Minimal pheomelanin",
        "RGB_Approximate": "(230, 200, 150)",
        "Hex_Code": "#E6C896",
        "Common_Names": "Platinum Blonde, Ice Blonde, Ash Blonde",
        "Percentage_Population": "<1% globally"
    },
    {
        "Level": 10,
        "Level_Name": "Lightest Blonde",
        "Description": "Palest natural blonde, almost white",
        "Underlying_Pigment": "Pale Yellow (minimal)",
        "Melanin_Type": "Very low melanin",
        "RGB_Approximate": "(245, 225, 180)",
        "Hex_Code": "#F5E1B4",
        "Common_Names": "White Blonde, Baby Blonde, Towhead",
        "Percentage_Population": "<0.5% globally"
    }
]

# =============================================================================
# UNDERTONE CLASSIFICATIONS
# =============================================================================

undertones = [
    {
        "Undertone": "Warm",
        "Description": "Golden, yellow, orange, or red undertones",
        "Characteristics": "Reflects warm light, appears sun-kissed",
        "Best_For": "Warm skin tones (yellow/golden undertones)",
        "Color_Examples": "Golden Blonde, Copper, Caramel, Honey Brown",
        "Hex_Example": "#D4A574"
    },
    {
        "Undertone": "Cool",
        "Description": "Ash, blue, violet, or silver undertones",
        "Characteristics": "Reflects cool light, no brassiness",
        "Best_For": "Cool skin tones (pink/blue undertones)",
        "Color_Examples": "Ash Blonde, Ash Brown, Platinum, Cool Black",
        "Hex_Example": "#8B7355"
    },
    {
        "Undertone": "Neutral",
        "Description": "Balanced mix of warm and cool tones",
        "Characteristics": "Versatile, neither warm nor cool dominant",
        "Best_For": "Neutral skin tones, universally flattering",
        "Color_Examples": "Natural Brown, Beige Blonde, Mushroom Brown",
        "Hex_Example": "#9C7A5B"
    }
]

# =============================================================================
# NATURAL HAIR COLOR SPECTRUM (Detailed)
# =============================================================================

natural_colors = [
    # BLACKS
    {"Category": "Black", "Shade_Name": "Jet Black", "Level": "1", "Undertone": "Neutral", "Hex": "#0A0A0A", "RGB": "10,10,10", "Description": "Purest black, blue-black sheen"},
    {"Category": "Black", "Shade_Name": "Blue-Black", "Level": "1", "Undertone": "Cool", "Hex": "#0D0D1A", "RGB": "13,13,26", "Description": "Black with blue undertones in light"},
    {"Category": "Black", "Shade_Name": "Soft Black", "Level": "1-2", "Undertone": "Warm", "Hex": "#1A1410", "RGB": "26,20,16", "Description": "Black with subtle brown warmth"},
    {"Category": "Black", "Shade_Name": "Raven", "Level": "1", "Undertone": "Cool", "Hex": "#0F0F0F", "RGB": "15,15,15", "Description": "Deep black with purple sheen"},
    
    # DARK BROWNS
    {"Category": "Dark Brown", "Shade_Name": "Espresso", "Level": "2", "Undertone": "Neutral", "Hex": "#231914", "RGB": "35,25,20", "Description": "Rich dark brown like espresso coffee"},
    {"Category": "Dark Brown", "Shade_Name": "Dark Chocolate", "Level": "2-3", "Undertone": "Warm", "Hex": "#2D1F1A", "RGB": "45,31,26", "Description": "Dark brown with red undertones"},
    {"Category": "Dark Brown", "Shade_Name": "Darkest Brown", "Level": "2", "Undertone": "Cool", "Hex": "#1E1614", "RGB": "30,22,20", "Description": "Nearly black brown, cool toned"},
    {"Category": "Dark Brown", "Shade_Name": "Coffee Bean", "Level": "2-3", "Undertone": "Neutral", "Hex": "#362820", "RGB": "54,40,32", "Description": "Deep brown like roasted coffee"},
    {"Category": "Dark Brown", "Shade_Name": "Cocoa", "Level": "3", "Undertone": "Warm", "Hex": "#3D2B1F", "RGB": "61,43,31", "Description": "Warm dark brown with red hints"},
    
    # MEDIUM BROWNS
    {"Category": "Medium Brown", "Shade_Name": "Chestnut", "Level": "4", "Undertone": "Warm", "Hex": "#5D3A1A", "RGB": "93,58,26", "Description": "Classic brown with red-gold undertones"},
    {"Category": "Medium Brown", "Shade_Name": "Milk Chocolate", "Level": "4", "Undertone": "Neutral", "Hex": "#5A3C28", "RGB": "90,60,40", "Description": "True medium brown"},
    {"Category": "Medium Brown", "Shade_Name": "Cinnamon", "Level": "4", "Undertone": "Warm", "Hex": "#6B4226", "RGB": "107,66,38", "Description": "Brown with strong copper undertones"},
    {"Category": "Medium Brown", "Shade_Name": "Walnut", "Level": "4", "Undertone": "Cool", "Hex": "#4A3728", "RGB": "74,55,40", "Description": "Cool-toned medium brown"},
    {"Category": "Medium Brown", "Shade_Name": "Mahogany", "Level": "4", "Undertone": "Warm", "Hex": "#5E3A32", "RGB": "94,58,50", "Description": "Brown with strong red undertones"},
    {"Category": "Medium Brown", "Shade_Name": "Auburn Brown", "Level": "4", "Undertone": "Warm", "Hex": "#6E3B2A", "RGB": "110,59,42", "Description": "Brown with prominent auburn tones"},
    
    # LIGHT BROWNS
    {"Category": "Light Brown", "Shade_Name": "Caramel", "Level": "5", "Undertone": "Warm", "Hex": "#8B5A2B", "RGB": "139,90,43", "Description": "Warm golden-brown like caramel candy"},
    {"Category": "Light Brown", "Shade_Name": "Toffee", "Level": "5", "Undertone": "Warm", "Hex": "#7B5544", "RGB": "123,85,68", "Description": "Rich warm brown with golden highlights"},
    {"Category": "Light Brown", "Shade_Name": "Light Chestnut", "Level": "5", "Undertone": "Warm", "Hex": "#8B6914", "RGB": "139,105,20", "Description": "Lighter chestnut with gold tones"},
    {"Category": "Light Brown", "Shade_Name": "Mushroom Brown", "Level": "5", "Undertone": "Cool", "Hex": "#7A6A5A", "RGB": "122,106,90", "Description": "Ashy light brown, muted tones"},
    {"Category": "Light Brown", "Shade_Name": "Hazelnut", "Level": "5", "Undertone": "Neutral", "Hex": "#8E7355", "RGB": "142,115,85", "Description": "Balanced light brown"},
    {"Category": "Light Brown", "Shade_Name": "Bronze", "Level": "5", "Undertone": "Warm", "Hex": "#8C6239", "RGB": "140,98,57", "Description": "Metallic warm brown"},
    
    # DARK BLONDES
    {"Category": "Dark Blonde", "Shade_Name": "Dirty Blonde", "Level": "6", "Undertone": "Neutral", "Hex": "#9A7B4F", "RGB": "154,123,79", "Description": "Brown-blonde transition color"},
    {"Category": "Dark Blonde", "Shade_Name": "Bronde", "Level": "6", "Undertone": "Warm", "Hex": "#A68B5B", "RGB": "166,139,91", "Description": "Perfect brown-blonde mix"},
    {"Category": "Dark Blonde", "Shade_Name": "Dark Honey", "Level": "6", "Undertone": "Warm", "Hex": "#B8860B", "RGB": "184,134,11", "Description": "Deep golden blonde"},
    {"Category": "Dark Blonde", "Shade_Name": "Dark Ash Blonde", "Level": "6", "Undertone": "Cool", "Hex": "#8B8378", "RGB": "139,131,120", "Description": "Cool-toned dark blonde"},
    {"Category": "Dark Blonde", "Shade_Name": "Caramel Blonde", "Level": "6", "Undertone": "Warm", "Hex": "#A67B5B", "RGB": "166,123,91", "Description": "Warm caramel-toned blonde"},
    
    # MEDIUM BLONDES
    {"Category": "Medium Blonde", "Shade_Name": "Golden Blonde", "Level": "7", "Undertone": "Warm", "Hex": "#DAA520", "RGB": "218,165,32", "Description": "Classic warm golden blonde"},
    {"Category": "Medium Blonde", "Shade_Name": "Honey Blonde", "Level": "7", "Undertone": "Warm", "Hex": "#C9A86C", "RGB": "201,168,108", "Description": "Sweet honey-toned blonde"},
    {"Category": "Medium Blonde", "Shade_Name": "Wheat Blonde", "Level": "7", "Undertone": "Neutral", "Hex": "#C4A35A", "RGB": "196,163,90", "Description": "Natural wheat-colored blonde"},
    {"Category": "Medium Blonde", "Shade_Name": "Medium Ash Blonde", "Level": "7", "Undertone": "Cool", "Hex": "#B5A191", "RGB": "181,161,145", "Description": "Cool ashy medium blonde"},
    {"Category": "Medium Blonde", "Shade_Name": "Butterscotch", "Level": "7", "Undertone": "Warm", "Hex": "#D4A76A", "RGB": "212,167,106", "Description": "Rich warm butterscotch blonde"},
    
    # LIGHT BLONDES
    {"Category": "Light Blonde", "Shade_Name": "Butter Blonde", "Level": "8", "Undertone": "Warm", "Hex": "#E6C88C", "RGB": "230,200,140", "Description": "Creamy butter-yellow blonde"},
    {"Category": "Light Blonde", "Shade_Name": "Champagne Blonde", "Level": "8", "Undertone": "Neutral", "Hex": "#E8D4B8", "RGB": "232,212,184", "Description": "Elegant champagne-toned blonde"},
    {"Category": "Light Blonde", "Shade_Name": "Sandy Blonde", "Level": "8", "Undertone": "Neutral", "Hex": "#D4B896", "RGB": "212,184,150", "Description": "Beach sand-colored blonde"},
    {"Category": "Light Blonde", "Shade_Name": "Light Ash Blonde", "Level": "8", "Undertone": "Cool", "Hex": "#C9C0B1", "RGB": "201,192,177", "Description": "Cool light ashy blonde"},
    {"Category": "Light Blonde", "Shade_Name": "Beige Blonde", "Level": "8", "Undertone": "Neutral", "Hex": "#D9C9A5", "RGB": "217,201,165", "Description": "Neutral beige-toned blonde"},
    
    # VERY LIGHT BLONDES
    {"Category": "Very Light Blonde", "Shade_Name": "Platinum Blonde", "Level": "9", "Undertone": "Cool", "Hex": "#E5E4E2", "RGB": "229,228,226", "Description": "Near-white cool blonde"},
    {"Category": "Very Light Blonde", "Shade_Name": "Ice Blonde", "Level": "9", "Undertone": "Cool", "Hex": "#F0EAE4", "RGB": "240,234,228", "Description": "Icy cool-toned blonde"},
    {"Category": "Very Light Blonde", "Shade_Name": "Pearl Blonde", "Level": "9", "Undertone": "Cool", "Hex": "#EAE0D5", "RGB": "234,224,213", "Description": "Pearlescent cool blonde"},
    {"Category": "Very Light Blonde", "Shade_Name": "Cream Blonde", "Level": "9", "Undertone": "Warm", "Hex": "#F5E6C8", "RGB": "245,230,200", "Description": "Creamy warm very light blonde"},
    
    # LIGHTEST BLONDES
    {"Category": "Lightest Blonde", "Shade_Name": "White Blonde", "Level": "10", "Undertone": "Neutral", "Hex": "#F5F5DC", "RGB": "245,245,220", "Description": "Near-white natural blonde"},
    {"Category": "Lightest Blonde", "Shade_Name": "Baby Blonde", "Level": "10", "Undertone": "Warm", "Hex": "#FAF0E6", "RGB": "250,240,230", "Description": "Pale baby-fine blonde"},
    {"Category": "Lightest Blonde", "Shade_Name": "Towhead", "Level": "10", "Undertone": "Neutral", "Hex": "#F8F4E3", "RGB": "248,244,227", "Description": "Extremely pale childhood blonde"},
    
    # REDS (Natural)
    {"Category": "Red", "Shade_Name": "Dark Auburn", "Level": "3-4", "Undertone": "Warm", "Hex": "#5C3317", "RGB": "92,51,23", "Description": "Deep red-brown"},
    {"Category": "Red", "Shade_Name": "Auburn", "Level": "4-5", "Undertone": "Warm", "Hex": "#8B4513", "RGB": "139,69,19", "Description": "Classic red-brown"},
    {"Category": "Red", "Shade_Name": "Copper", "Level": "5-6", "Undertone": "Warm", "Hex": "#B87333", "RGB": "184,115,51", "Description": "Bright copper-orange red"},
    {"Category": "Red", "Shade_Name": "Ginger", "Level": "6-7", "Undertone": "Warm", "Hex": "#B06500", "RGB": "176,101,0", "Description": "Orange-toned natural red"},
    {"Category": "Red", "Shade_Name": "Strawberry Blonde", "Level": "7-8", "Undertone": "Warm", "Hex": "#D4A484", "RGB": "212,164,132", "Description": "Blonde with pink-red tones"},
    {"Category": "Red", "Shade_Name": "Titian", "Level": "5-6", "Undertone": "Warm", "Hex": "#CD5C5C", "RGB": "205,92,92", "Description": "Brownish-orange red"},
    
    # GRAYS & WHITES
    {"Category": "Gray/White", "Shade_Name": "Salt & Pepper", "Level": "Variable", "Undertone": "Neutral", "Hex": "#808080", "RGB": "128,128,128", "Description": "Mix of gray and natural color"},
    {"Category": "Gray/White", "Shade_Name": "Silver", "Level": "9-10", "Undertone": "Cool", "Hex": "#C0C0C0", "RGB": "192,192,192", "Description": "Metallic gray"},
    {"Category": "Gray/White", "Shade_Name": "Steel Gray", "Level": "7-8", "Undertone": "Cool", "Hex": "#71797E", "RGB": "113,121,126", "Description": "Blue-toned gray"},
    {"Category": "Gray/White", "Shade_Name": "Charcoal", "Level": "3-4", "Undertone": "Cool", "Hex": "#36454F", "RGB": "54,69,79", "Description": "Dark gray"},
    {"Category": "Gray/White", "Shade_Name": "Pure White", "Level": "10+", "Undertone": "Neutral", "Hex": "#FFFFFF", "RGB": "255,255,255", "Description": "Complete absence of pigment"},
    {"Category": "Gray/White", "Shade_Name": "Snow White", "Level": "10+", "Undertone": "Cool", "Hex": "#FFFAFA", "RGB": "255,250,250", "Description": "Bright white with cool tone"},
]

# =============================================================================
# ARTIFICIAL/FANTASY COLORS
# =============================================================================

fantasy_colors = [
    # REDS (Artificial)
    {"Category": "Fashion Red", "Shade_Name": "Bright Red", "Hex": "#FF0000", "RGB": "255,0,0", "Description": "Pure bright red"},
    {"Category": "Fashion Red", "Shade_Name": "Cherry Red", "Hex": "#DE3163", "RGB": "222,49,99", "Description": "Blue-toned red"},
    {"Category": "Fashion Red", "Shade_Name": "Burgundy", "Hex": "#800020", "RGB": "128,0,32", "Description": "Deep wine red"},
    {"Category": "Fashion Red", "Shade_Name": "Crimson", "Hex": "#DC143C", "RGB": "220,20,60", "Description": "Deep red with blue undertones"},
    {"Category": "Fashion Red", "Shade_Name": "Fire Engine Red", "Hex": "#CE2029", "RGB": "206,32,41", "Description": "Bright true red"},
    
    # PINKS
    {"Category": "Pink", "Shade_Name": "Hot Pink", "Hex": "#FF69B4", "RGB": "255,105,180", "Description": "Vibrant bright pink"},
    {"Category": "Pink", "Shade_Name": "Pastel Pink", "Hex": "#FFD1DC", "RGB": "255,209,220", "Description": "Soft pale pink"},
    {"Category": "Pink", "Shade_Name": "Rose Gold", "Hex": "#B76E79", "RGB": "183,110,121", "Description": "Metallic pink-gold"},
    {"Category": "Pink", "Shade_Name": "Magenta", "Hex": "#FF00FF", "RGB": "255,0,255", "Description": "Blue-red pink"},
    {"Category": "Pink", "Shade_Name": "Dusty Rose", "Hex": "#DCAE96", "RGB": "220,174,150", "Description": "Muted mauve-pink"},
    {"Category": "Pink", "Shade_Name": "Bubblegum", "Hex": "#FFC1CC", "RGB": "255,193,204", "Description": "Bright candy pink"},
    
    # PURPLES
    {"Category": "Purple", "Shade_Name": "Violet", "Hex": "#8B00FF", "RGB": "139,0,255", "Description": "True violet"},
    {"Category": "Purple", "Shade_Name": "Lavender", "Hex": "#E6E6FA", "RGB": "230,230,250", "Description": "Pale purple"},
    {"Category": "Purple", "Shade_Name": "Plum", "Hex": "#8E4585", "RGB": "142,69,133", "Description": "Deep purple-red"},
    {"Category": "Purple", "Shade_Name": "Grape", "Hex": "#6F2DA8", "RGB": "111,45,168", "Description": "Rich grape purple"},
    {"Category": "Purple", "Shade_Name": "Lilac", "Hex": "#C8A2C8", "RGB": "200,162,200", "Description": "Soft pastel purple"},
    {"Category": "Purple", "Shade_Name": "Eggplant", "Hex": "#614051", "RGB": "97,64,81", "Description": "Dark purple-brown"},
    
    # BLUES
    {"Category": "Blue", "Shade_Name": "Electric Blue", "Hex": "#7DF9FF", "RGB": "125,249,255", "Description": "Bright cyan-blue"},
    {"Category": "Blue", "Shade_Name": "Navy Blue", "Hex": "#000080", "RGB": "0,0,128", "Description": "Deep dark blue"},
    {"Category": "Blue", "Shade_Name": "Sky Blue", "Hex": "#87CEEB", "RGB": "135,206,235", "Description": "Light sky blue"},
    {"Category": "Blue", "Shade_Name": "Teal", "Hex": "#008080", "RGB": "0,128,128", "Description": "Blue-green"},
    {"Category": "Blue", "Shade_Name": "Cobalt", "Hex": "#0047AB", "RGB": "0,71,171", "Description": "Deep true blue"},
    {"Category": "Blue", "Shade_Name": "Pastel Blue", "Hex": "#AEC6CF", "RGB": "174,198,207", "Description": "Soft pale blue"},
    {"Category": "Blue", "Shade_Name": "Denim", "Hex": "#1560BD", "RGB": "21,96,189", "Description": "Medium denim blue"},
    
    # GREENS
    {"Category": "Green", "Shade_Name": "Emerald", "Hex": "#50C878", "RGB": "80,200,120", "Description": "Bright jewel green"},
    {"Category": "Green", "Shade_Name": "Forest Green", "Hex": "#228B22", "RGB": "34,139,34", "Description": "Deep forest green"},
    {"Category": "Green", "Shade_Name": "Mint", "Hex": "#98FF98", "RGB": "152,255,152", "Description": "Pale mint green"},
    {"Category": "Green", "Shade_Name": "Olive", "Hex": "#808000", "RGB": "128,128,0", "Description": "Yellow-green olive"},
    {"Category": "Green", "Shade_Name": "Neon Green", "Hex": "#39FF14", "RGB": "57,255,20", "Description": "Bright fluorescent green"},
    {"Category": "Green", "Shade_Name": "Seafoam", "Hex": "#71EEB8", "RGB": "113,238,184", "Description": "Blue-green seafoam"},
    
    # ORANGES
    {"Category": "Orange", "Shade_Name": "Bright Orange", "Hex": "#FF7F00", "RGB": "255,127,0", "Description": "Pure bright orange"},
    {"Category": "Orange", "Shade_Name": "Peach", "Hex": "#FFCBA4", "RGB": "255,203,164", "Description": "Soft peach orange"},
    {"Category": "Orange", "Shade_Name": "Tangerine", "Hex": "#FF9966", "RGB": "255,153,102", "Description": "Red-orange tangerine"},
    {"Category": "Orange", "Shade_Name": "Coral", "Hex": "#FF7F50", "RGB": "255,127,80", "Description": "Pink-orange coral"},
    {"Category": "Orange", "Shade_Name": "Pumpkin", "Hex": "#FF7518", "RGB": "255,117,24", "Description": "Deep pumpkin orange"},
    
    # YELLOWS
    {"Category": "Yellow", "Shade_Name": "Bright Yellow", "Hex": "#FFFF00", "RGB": "255,255,0", "Description": "Pure bright yellow"},
    {"Category": "Yellow", "Shade_Name": "Pastel Yellow", "Hex": "#FDFD96", "RGB": "253,253,150", "Description": "Soft pale yellow"},
    {"Category": "Yellow", "Shade_Name": "Gold", "Hex": "#FFD700", "RGB": "255,215,0", "Description": "Metallic gold"},
    {"Category": "Yellow", "Shade_Name": "Mustard", "Hex": "#FFDB58", "RGB": "255,219,88", "Description": "Yellow-brown mustard"},
    
    # MULTI/SPECIAL
    {"Category": "Special", "Shade_Name": "Rainbow", "Hex": "Multi", "RGB": "Multi", "Description": "Multiple colors in gradient"},
    {"Category": "Special", "Shade_Name": "Oil Slick", "Hex": "Multi", "RGB": "Multi", "Description": "Dark base with rainbow sheen"},
    {"Category": "Special", "Shade_Name": "Holographic", "Hex": "Multi", "RGB": "Multi", "Description": "Color-shifting iridescent"},
    {"Category": "Special", "Shade_Name": "Ombre", "Hex": "Gradient", "RGB": "Gradient", "Description": "Gradual color transition"},
    {"Category": "Special", "Shade_Name": "Balayage", "Hex": "Gradient", "RGB": "Gradient", "Description": "Hand-painted highlights"},
]

# =============================================================================
# COLOR TREATMENT TYPES
# =============================================================================

treatment_types = [
    {"Treatment": "Virgin/Natural", "Description": "No chemical color treatment", "Permanence": "N/A", "Damage_Level": "None"},
    {"Treatment": "Permanent Color", "Description": "Penetrates cortex, lifts and deposits color", "Permanence": "Until grow-out", "Damage_Level": "Medium-High"},
    {"Treatment": "Demi-Permanent", "Description": "Deposits color without lifting", "Permanence": "20-28 washes", "Damage_Level": "Low-Medium"},
    {"Treatment": "Semi-Permanent", "Description": "Coats hair shaft only", "Permanence": "8-12 washes", "Damage_Level": "None-Low"},
    {"Treatment": "Temporary/Wash-Out", "Description": "Surface coating only", "Permanence": "1-2 washes", "Damage_Level": "None"},
    {"Treatment": "Bleach/Lightener", "Description": "Removes melanin from hair", "Permanence": "Permanent", "Damage_Level": "High"},
    {"Treatment": "Highlights", "Description": "Lightened sections throughout", "Permanence": "Until grow-out", "Damage_Level": "Medium-High"},
    {"Treatment": "Lowlights", "Description": "Darker sections added for depth", "Permanence": "Until grow-out", "Damage_Level": "Low-Medium"},
    {"Treatment": "Balayage", "Description": "Hand-painted highlights for natural gradient", "Permanence": "Until grow-out", "Damage_Level": "Medium"},
    {"Treatment": "Ombre", "Description": "Gradual dark-to-light transition", "Permanence": "Until grow-out", "Damage_Level": "Medium-High"},
    {"Treatment": "Color Melting", "Description": "Seamless blend of multiple colors", "Permanence": "Until grow-out", "Damage_Level": "Medium"},
    {"Treatment": "Gloss/Glaze", "Description": "Adds shine and tonal enhancement", "Permanence": "4-6 weeks", "Damage_Level": "None-Low"},
    {"Treatment": "Toner", "Description": "Neutralizes unwanted tones", "Permanence": "4-8 weeks", "Damage_Level": "Low"},
    {"Treatment": "Root Touch-Up", "Description": "Color applied to regrowth only", "Permanence": "Until grow-out", "Damage_Level": "Low-Medium"},
    {"Treatment": "Color Correction", "Description": "Fixing previous color mistakes", "Permanence": "Varies", "Damage_Level": "High"},
]

# =============================================================================
# CREATE EXCEL FILE
# =============================================================================

def create_color_taxonomy_excel():
    # Create DataFrames
    df_levels = pd.DataFrame(color_levels)
    df_undertones = pd.DataFrame(undertones)
    df_natural = pd.DataFrame(natural_colors)
    df_fantasy = pd.DataFrame(fantasy_colors)
    df_treatments = pd.DataFrame(treatment_types)
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter('/home/ubuntu/hair_taxonomy/excel/hair_color_taxonomy.xlsx', engine='openpyxl') as writer:
        df_levels.to_excel(writer, sheet_name='Color Levels (1-10)', index=False)
        df_undertones.to_excel(writer, sheet_name='Undertones', index=False)
        df_natural.to_excel(writer, sheet_name='Natural Colors', index=False)
        df_fantasy.to_excel(writer, sheet_name='Fantasy Colors', index=False)
        df_treatments.to_excel(writer, sheet_name='Treatment Types', index=False)
        
        # Auto-adjust column widths
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print("Hair Color Taxonomy Excel created successfully!")

if __name__ == "__main__":
    create_color_taxonomy_excel()
