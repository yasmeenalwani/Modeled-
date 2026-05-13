#!/usr/bin/env python3
"""
Creates comprehensive Excel files summarizing all free datasets
for the Modeled Hair Engine.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Dataset Information ---
datasets = [
    {
        "Dataset Name": "CelebA",
        "Size (Images)": "202,599",
        "Resolution": "178×218 (aligned)",
        "Hair Features": "Color (4 types), Type (2), Bangs, Bald",
        "Segmentation Masks": "No",
        "License": "Non-Commercial Research",
        "Commercial Use": "NO",
        "Download Source": "Google Drive / Baidu Drive",
        "URL": "https://mmlab.ie.cuhk.edu.hk/projects/CelebA.html",
        "Key Advantage": "Largest face attribute dataset",
        "Key Limitation": "Limited hair granularity, celebrity bias",
        "Recommended Use": "MVP prototyping, transfer learning"
    },
    {
        "Dataset Name": "CelebAMask-HQ",
        "Size (Images)": "30,000",
        "Resolution": "512×512",
        "Hair Features": "Hair segmentation masks",
        "Segmentation Masks": "Yes (19 classes)",
        "License": "Non-Commercial Research",
        "Commercial Use": "NO",
        "Download Source": "Google Drive / Baidu Drive",
        "URL": "https://github.com/switchablenorms/CelebAMask-HQ",
        "Key Advantage": "High-quality hair segmentation masks",
        "Key Limitation": "Subset of CelebA, non-commercial",
        "Recommended Use": "Training hair segmentation models"
    },
    {
        "Dataset Name": "Figaro1k",
        "Size (Images)": "1,050",
        "Resolution": "Varies (normalized to 227×227)",
        "Hair Features": "7 hairstyle classes, segmentation",
        "Segmentation Masks": "Yes",
        "License": "Fair Use (Non-Profit)",
        "Commercial Use": "NO",
        "Download Source": "OSF (Open Science Framework)",
        "URL": "https://osf.io/wg5u2/",
        "Key Advantage": "Explicit curl pattern labels",
        "Key Limitation": "Small size, web-scraped images",
        "Recommended Use": "Curl pattern classification training"
    },
    {
        "Dataset Name": "LFW Hair (Parts)",
        "Size (Images)": "~2,000",
        "Resolution": "250×250",
        "Hair Features": "Hair segmentation masks",
        "Segmentation Masks": "Yes",
        "License": "Fair Use (Gray Area)",
        "Commercial Use": "NOT RECOMMENDED",
        "Download Source": "Academic sources",
        "URL": "http://vis-www.cs.umass.edu/lfw/",
        "Key Advantage": "Real-world face images",
        "Key Limitation": "Legal uncertainty for commercial use",
        "Recommended Use": "Academic research only"
    },
    {
        "Dataset Name": "UTKFace",
        "Size (Images)": "~20,000",
        "Resolution": "200×200",
        "Hair Features": "Age, Gender, Ethnicity (indirect)",
        "Segmentation Masks": "No",
        "License": "Non-Commercial Research",
        "Commercial Use": "NO",
        "Download Source": "Kaggle",
        "URL": "https://susanqq.github.io/UTKFace/",
        "Key Advantage": "Demographic diversity, age range",
        "Key Limitation": "No direct hair labels",
        "Recommended Use": "Demographic correlation analysis"
    },
    {
        "Dataset Name": "Black Hair Detection",
        "Size (Images)": "704",
        "Resolution": "Varies",
        "Hair Features": "10 inclusive hairstyle classes",
        "Segmentation Masks": "Bounding boxes",
        "License": "CC BY 4.0",
        "Commercial Use": "YES (with attribution)",
        "Download Source": "Roboflow Universe",
        "URL": "https://universe.roboflow.com/aishas-workspace/black-hair-detection",
        "Key Advantage": "Inclusive styles, commercial license",
        "Key Limitation": "Smaller size",
        "Recommended Use": "Commercial product training"
    },
    {
        "Dataset Name": "FairFace",
        "Size (Images)": "108,501",
        "Resolution": "224×224",
        "Hair Features": "Balanced demographics",
        "Segmentation Masks": "No",
        "License": "Apache 2.0",
        "Commercial Use": "YES",
        "Download Source": "GitHub / Hugging Face",
        "URL": "https://github.com/joojs/fairface",
        "Key Advantage": "Racially balanced, commercial license",
        "Key Limitation": "No direct hair labels",
        "Recommended Use": "Fairness testing, commercial training"
    },
    {
        "Dataset Name": "FFHQ",
        "Size (Images)": "70,000",
        "Resolution": "1024×1024",
        "Hair Features": "High-resolution faces",
        "Segmentation Masks": "No (but available separately)",
        "License": "Varies (Creative Commons)",
        "Commercial Use": "VARIES (per image)",
        "Download Source": "Kaggle / Archive.org / NVIDIA",
        "URL": "https://github.com/NVlabs/ffhq-dataset",
        "Key Advantage": "Highest quality, diverse",
        "Key Limitation": "License filtering required",
        "Recommended Use": "GAN training, high-res analysis"
    },
    {
        "Dataset Name": "LIP (Look Into Person)",
        "Size (Images)": "50,462",
        "Resolution": "Varies",
        "Hair Features": "Hair as 1 of 20 body parts",
        "Segmentation Masks": "Yes (20 classes)",
        "License": "Research Use",
        "Commercial Use": "NO",
        "Download Source": "SYSU-HCP Lab",
        "URL": "https://www.sysu-hcp.net/resources/datasets/",
        "Key Advantage": "Full-body parsing with hair",
        "Key Limitation": "Not face-focused",
        "Recommended Use": "Full-body hair detection"
    },
    {
        "Dataset Name": "CIHP",
        "Size (Images)": "38,280",
        "Resolution": "Varies",
        "Hair Features": "Multi-person hair parsing",
        "Segmentation Masks": "Yes (instance-level)",
        "License": "Research Use",
        "Commercial Use": "NO",
        "Download Source": "LIP Challenge",
        "URL": "Via LIP dataset page",
        "Key Advantage": "Multi-person scenes",
        "Key Limitation": "Complex, research-only",
        "Recommended Use": "Crowd scene hair analysis"
    },
]

# --- Hair Feature Details ---
hair_features = [
    {
        "Dataset": "CelebA",
        "Black Hair": "Yes (binary)",
        "Brown Hair": "Yes (binary)",
        "Blonde Hair": "Yes (binary)",
        "Gray Hair": "Yes (binary)",
        "Red Hair": "No",
        "Straight Hair": "Yes (binary)",
        "Wavy Hair": "Yes (binary)",
        "Curly Hair": "No",
        "Coily/Kinky": "No",
        "Braids": "No",
        "Locs": "No",
        "Bangs": "Yes (binary)",
        "Bald": "Yes (binary)",
        "Hair Mask": "No"
    },
    {
        "Dataset": "CelebAMask-HQ",
        "Black Hair": "No",
        "Brown Hair": "No",
        "Blonde Hair": "No",
        "Gray Hair": "No",
        "Red Hair": "No",
        "Straight Hair": "No",
        "Wavy Hair": "No",
        "Curly Hair": "No",
        "Coily/Kinky": "No",
        "Braids": "No",
        "Locs": "No",
        "Bangs": "No",
        "Bald": "No",
        "Hair Mask": "Yes (pixel-level)"
    },
    {
        "Dataset": "Figaro1k",
        "Black Hair": "No",
        "Brown Hair": "No",
        "Blonde Hair": "No",
        "Gray Hair": "No",
        "Red Hair": "No",
        "Straight Hair": "Yes (class)",
        "Wavy Hair": "Yes (class)",
        "Curly Hair": "Yes (class)",
        "Coily/Kinky": "Yes (class)",
        "Braids": "Yes (class)",
        "Locs": "Yes (class)",
        "Bangs": "No",
        "Bald": "No",
        "Hair Mask": "Yes (pixel-level)"
    },
    {
        "Dataset": "Black Hair (Roboflow)",
        "Black Hair": "No",
        "Brown Hair": "No",
        "Blonde Hair": "No",
        "Gray Hair": "No",
        "Red Hair": "No",
        "Straight Hair": "No",
        "Wavy Hair": "No",
        "Curly Hair": "No",
        "Coily/Kinky": "No",
        "Braids": "Yes (class)",
        "Locs": "Yes (class)",
        "Bangs": "No",
        "Bald": "No",
        "Hair Mask": "Bounding box"
    },
]

# --- Challenges Summary ---
challenges = [
    {
        "Challenge Category": "Licensing",
        "Challenge": "Non-Commercial Restrictions",
        "Impact": "HIGH",
        "Description": "Most academic datasets (CelebA, Figaro1k, UTKFace) prohibit commercial use",
        "Mitigation Strategy": "Use for R&D only; build proprietary dataset for production"
    },
    {
        "Challenge Category": "Licensing",
        "Challenge": "Image Copyright",
        "Impact": "HIGH",
        "Description": "Web-scraped images retain original copyright; legal risk for commercial use",
        "Mitigation Strategy": "Prioritize CC-licensed datasets (Roboflow, FairFace)"
    },
    {
        "Challenge Category": "Data Quality",
        "Challenge": "Annotation Inconsistency",
        "Impact": "MEDIUM",
        "Description": "Different datasets use different label definitions and granularity",
        "Mitigation Strategy": "Create master taxonomy; re-label data to unified schema"
    },
    {
        "Challenge Category": "Data Quality",
        "Challenge": "Label Noise",
        "Impact": "MEDIUM",
        "Description": "Binary attributes in CelebA are often inaccurate or subjective",
        "Mitigation Strategy": "Use consensus labeling; implement data cleaning pipeline"
    },
    {
        "Challenge Category": "Bias",
        "Challenge": "Demographic Underrepresentation",
        "Impact": "HIGH",
        "Description": "Celebrity datasets skew towards certain demographics and hair types",
        "Mitigation Strategy": "Supplement with FairFace and Black Hair datasets"
    },
    {
        "Challenge Category": "Bias",
        "Challenge": "Curl Pattern Imbalance",
        "Impact": "HIGH",
        "Description": "Type 4 (coily/kinky) hair is underrepresented in most datasets",
        "Mitigation Strategy": "Active data sourcing; use inclusive datasets like Roboflow"
    },
    {
        "Challenge Category": "Technical",
        "Challenge": "Environmental Variability",
        "Impact": "MEDIUM",
        "Description": "Lighting, hair state (wet/dry), and image quality vary widely",
        "Mitigation Strategy": "Aggressive data augmentation; capture contextual metadata"
    },
    {
        "Challenge Category": "Technical",
        "Challenge": "Fine-Grained Classification",
        "Impact": "MEDIUM",
        "Description": "Distinguishing between similar shades/patterns requires more data",
        "Mitigation Strategy": "Hierarchical classification; metric learning"
    },
    {
        "Challenge Category": "Business",
        "Challenge": "Vendor Lock-in",
        "Impact": "LOW",
        "Description": "Over-reliance on AWS Rekognition can limit flexibility",
        "Mitigation Strategy": "Parallel development of proprietary models"
    },
    {
        "Challenge Category": "Business",
        "Challenge": "Cost at Scale",
        "Impact": "MEDIUM",
        "Description": "Cloud ML services become expensive at high inference volumes",
        "Mitigation Strategy": "Cost-benefit analysis; migrate to self-hosted at scale"
    },
]

# --- Create Excel File ---
def create_excel():
    output_path = "/home/ubuntu/project_guides/free_datasets_comprehensive_guide.xlsx"
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Sheet 1: Master Dataset Summary
        df_datasets = pd.DataFrame(datasets)
        df_datasets.to_excel(writer, sheet_name='Dataset Summary', index=False)
        
        # Sheet 2: Hair Feature Matrix
        df_features = pd.DataFrame(hair_features)
        df_features.to_excel(writer, sheet_name='Hair Feature Matrix', index=False)
        
        # Sheet 3: Challenges & Considerations
        df_challenges = pd.DataFrame(challenges)
        df_challenges.to_excel(writer, sheet_name='Challenges', index=False)
        
        # Sheet 4: Download Instructions
        download_data = [
            {"Dataset": "CelebA", "Method": "Google Drive", "Command/URL": "https://drive.google.com/drive/folders/0B7EVK8r0v71pWEZsZE9oNnFzTm8"},
            {"Dataset": "CelebAMask-HQ", "Method": "Google Drive", "Command/URL": "https://drive.google.com/open?id=1badu11NqxGepkS-ang-LvDPKQyxC4W57"},
            {"Dataset": "Figaro1k", "Method": "OSF", "Command/URL": "https://osf.io/wg5u2/"},
            {"Dataset": "Black Hair", "Method": "Roboflow", "Command/URL": "pip install roboflow; use API"},
            {"Dataset": "FairFace", "Method": "Hugging Face", "Command/URL": "datasets.load_dataset('HuggingFaceM4/FairFace')"},
            {"Dataset": "UTKFace", "Method": "Kaggle", "Command/URL": "kaggle datasets download -d jangedoo/utkface-new"},
            {"Dataset": "FFHQ", "Method": "Kaggle", "Command/URL": "kaggle datasets download -d arnaud58/flickrfaceshq-dataset-ffhq"},
            {"Dataset": "LIP", "Method": "Official Site", "Command/URL": "https://www.sysu-hcp.net/resources/datasets/"},
        ]
        df_download = pd.DataFrame(download_data)
        df_download.to_excel(writer, sheet_name='Download Instructions', index=False)
        
        # Sheet 5: Recommended Strategy
        strategy_data = [
            {"Phase": "Phase 1: MVP", "Primary Datasets": "CelebA, CelebAMask-HQ, Figaro1k", "Purpose": "Initial model training and testing", "License OK for Phase": "Yes (internal R&D)"},
            {"Phase": "Phase 2: Fairness Testing", "Primary Datasets": "FairFace, Black Hair (Roboflow)", "Purpose": "Evaluate demographic performance", "License OK for Phase": "Yes"},
            {"Phase": "Phase 3: Transfer Learning", "Primary Datasets": "CelebAMask-HQ, CelebA", "Purpose": "Pre-train models for feature extraction", "License OK for Phase": "Yes (internal)"},
            {"Phase": "Phase 4: Production", "Primary Datasets": "Proprietary + Black Hair + FairFace", "Purpose": "Train commercial product", "License OK for Phase": "Yes (commercial)"},
        ]
        df_strategy = pd.DataFrame(strategy_data)
        df_strategy.to_excel(writer, sheet_name='Recommended Strategy', index=False)
    
    print(f"Excel file created: {output_path}")
    
    # Apply formatting
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    
    # Style settings
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    wb.save(output_path)
    print("Formatting applied successfully!")

if __name__ == "__main__":
    create_excel()
